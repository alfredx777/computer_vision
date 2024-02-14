import cv2
import os
import openpyxl
from openpyxl import Workbook

# Load the trained facial recognition model
model = cv2.face.LBPHFaceRecognizer_create()
model.read('face_recognition_model.yml')

# Load the pre-trained Haar Cascade face detector
face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')

# Initialize the webcam
cap = cv2.VideoCapture(0)

# Define a dictionary to map label numbers to names
label_names = {
    0: 'Person 1',
    1: 'Person 2',
    # Add more entries as needed
}

# Load or create the Excel workbook and sheet
try:
    workbook = openpyxl.load_workbook('face_recognition_records.xlsx')
except FileNotFoundError:
    workbook = Workbook()
sheet = workbook.active
if 'Records' not in workbook.sheetnames:
    sheet.title = 'Records'
else:
    sheet = workbook['Records']

# Set to keep track of recognized faces
recognized_faces = set()

# Create the directory for unknown faces if it doesn't exist
unknown_dir = 'the_unknown'
if not os.path.exists(unknown_dir):
    os.makedirs(unknown_dir)

while True:
    # Capture a frame from the webcam
    ret, frame = cap.read()

    # Convert the frame to grayscale
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)

    # Detect faces in the grayscale frame
    faces = face_cascade.detectMultiScale(gray, scaleFactor=1.1, minNeighbors=5, minSize=(30, 30))

    # Loop through the detected faces
    for (x, y, w, h) in faces:
        # Extract the face region
        face_roi = gray[y:y+h, x:x+w]

        # Recognize the face using the trained model
        label, confidence = model.predict(face_roi)

        # Display the name and confidence if the label exists in the dictionary
        if label in label_names:
            name = label_names[label]
        else:
            name = 'Unknown'
            unknown_image_path = os.path.join(unknown_dir, f'unknown_{len(os.listdir(unknown_dir)) + 1}.jpg')
            cv2.imwrite(unknown_image_path, frame[y:y+h, x:x+w])

        text = f'Name: {name}, Confidence: {confidence:.2f}'
        cv2.putText(frame, text, (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 0), 2)

        # Record the name only if it has not been recorded before
        if name not in recognized_faces:
            recognized_faces.add(name)
            sheet.append([name])
            workbook.save('face_recognition_records.xlsx')

        # Draw a rectangle around the face
        cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 2)

    # Display the frame with detected faces and recognition results
    cv2.imshow('Face Recognition', frame)

    # Wait for key press
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

# Release the webcam and close OpenCV windows
cap.release()
cv2.destroyAllWindows()
